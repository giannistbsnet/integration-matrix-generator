from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .components import Component

DESCRIPTIONS_SHEET = "Descriptions"
MATRIX_SHEET = "Integration Matrix"

DIRECTION_OPTIONS = ["↔", "↗", "↙"]
INTERFACE_OPTIONS = [
    "REST API",
    "GraphQL",
    "gRPC",
    "Kafka",
    "MQTT",
    "WebSocket",
    "SPARQL endpoint",
    "NGSI-LD",
    "Eclipse Dataspace Connector (EDC)",
    "AAS / Eclipse BaSyx",
    "File / Batch",
    "Other",
]

HEADER_FILL = "D8D8D8"
COMPONENT_FILL = "D9E0FB"
SUBHEADER_FILL = "FDBF11"
BLACK_FILL = "000000"
MIRROR_FILL = "F0F0F0"
TEXT_DARK = "000000"
TEXT_BLUE = "262668"
TEXT_GRAY = "434343"

THIN_GRAY = Side(style="thin", color="B7B7B7")
THIN_BLACK = Side(style="thin", color="000000")

START_ROW = 3
START_COL = 4


def create_workbook(components: Sequence[Component]) -> Workbook:
    if not components:
        raise ValueError("At least one component is required.")

    workbook = Workbook()
    descriptions = workbook.active
    descriptions.title = DESCRIPTIONS_SHEET
    matrix = workbook.create_sheet(MATRIX_SHEET, 0)

    _build_descriptions_sheet(descriptions, components)
    _build_matrix_sheet(matrix, components)

    workbook.active = workbook.sheetnames.index(MATRIX_SHEET)
    return workbook


def save_workbook(workbook: Workbook, path: str | Path) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _build_descriptions_sheet(sheet, components: Sequence[Component]) -> None:
    headers = [
        "Partner",
        "Task",
        "Component",
        "Description of Component",
        "Description of Interface",
    ]
    sheet.append(headers)
    for component in components:
        sheet.append([component.partner, component.task, component.component, None, None])

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 14.5
    sheet.column_dimensions["B"].width = 11.5
    sheet.column_dimensions["C"].width = 35.13
    sheet.column_dimensions["D"].width = 37.13
    sheet.column_dimensions["E"].width = 25.88
    sheet.row_dimensions[1].height = 22.5

    table_ref = f"A1:E{len(components) + 1}"
    table = Table(displayName="Descriptions", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color=TEXT_DARK)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2, max_row=len(components) + 1, min_col=1, max_col=5):
        sheet.row_dimensions[row[0].row].height = 45
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column >= 4)

    comment_text = (
        "1. What protocol/technology do you expose?\n\n"
        "2. What data do you send/receive and in what format?\n\n"
        "3. What does a consumer need to connect to you?"
    )
    sheet["E1"].comment = Comment(comment_text, "Integration Matrix Generator")


def _build_matrix_sheet(sheet, components: Sequence[Component]) -> None:
    count = len(components)
    last_col = START_COL + (count * 2) - 1
    last_row = START_ROW + count - 1

    sheet.freeze_panes = "D3"
    sheet.row_dimensions[1].height = 22.5
    sheet.row_dimensions[2].height = 30.75
    for row in range(START_ROW, last_row + 1):
        sheet.row_dimensions[row].height = 27.75

    _set_matrix_dimensions(sheet, last_col)
    _write_matrix_headers(sheet, components)
    _write_matrix_rows(sheet, components)
    _write_matrix_cells(sheet, count)
    _add_matrix_validations(sheet, count)


def _set_matrix_dimensions(sheet, last_col: int) -> None:
    sheet.column_dimensions["A"].width = 13.38
    sheet.column_dimensions["B"].width = 8.38
    sheet.column_dimensions["C"].width = 36.63
    for col in range(START_COL, last_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 11.63 if (col - START_COL) % 2 == 0 else 13.0


def _write_matrix_headers(sheet, components: Sequence[Component]) -> None:
    base_headers = ["Partner", "Task", "Component"]
    for col, label in enumerate(base_headers, start=1):
        cell = sheet.cell(row=1, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Arial", size=13, bold=True, color=TEXT_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=THIN_BLACK, bottom=THIN_BLACK, left=THIN_BLACK, right=THIN_BLACK)
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    for index, component in enumerate(components):
        direction_col = START_COL + (index * 2)
        interface_col = direction_col + 1
        sheet.merge_cells(start_row=1, start_column=direction_col, end_row=1, end_column=interface_col)

        header = sheet.cell(row=1, column=direction_col, value=component.component)
        header.fill = PatternFill("solid", fgColor=COMPONENT_FILL)
        header.font = Font(name="Arial", size=12, color=TEXT_DARK)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = Border(top=THIN_BLACK, right=THIN_BLACK)

        for col, label in [(direction_col, "Data Flow Direction"), (interface_col, "Interface")]:
            subheader = sheet.cell(row=2, column=col, value=label)
            subheader.fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
            subheader.font = Font(name="Arial", size=12, bold=True, color=TEXT_BLUE)
            subheader.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            subheader.border = Border(bottom=THIN_BLACK, right=THIN_BLACK if col == interface_col else None)


def _write_matrix_rows(sheet, components: Sequence[Component]) -> None:
    for index, component in enumerate(components):
        row = START_ROW + index
        values = [component.partner, component.task, component.component]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL if col < 3 else COMPONENT_FILL)
            cell.font = Font(
                name="Roboto" if col < 3 else "Arial",
                size=11,
                bold=col < 3,
                color=TEXT_GRAY if col < 3 else TEXT_DARK,
            )
            cell.alignment = Alignment(horizontal="center" if col < 3 else None, vertical="center", wrap_text=col == 3)
            cell.border = Border(left=THIN_BLACK if col == 1 else None, right=THIN_BLACK if col == 3 else None)


def _write_matrix_cells(sheet, count: int) -> None:
    for row_index in range(count):
        row = START_ROW + row_index
        for col_index in range(count):
            direction_col = START_COL + (col_index * 2)
            interface_col = direction_col + 1
            direction_cell = sheet.cell(row=row, column=direction_col)
            interface_cell = sheet.cell(row=row, column=interface_col)

            if row_index == col_index:
                _style_matrix_pair(direction_cell, interface_cell, fill=BLACK_FILL)
                continue

            if row_index < col_index:
                mirror_row = START_ROW + col_index
                mirror_direction_col = START_COL + (row_index * 2)
                mirror_interface_col = mirror_direction_col + 1
                direction_ref = f"{get_column_letter(mirror_direction_col)}{mirror_row}"
                interface_ref = f"{get_column_letter(mirror_interface_col)}{mirror_row}"
                direction_cell.value = f'=IF({direction_ref}="↗","↙",IF({direction_ref}="↙","↗",{direction_ref}))'
                interface_cell.value = f"={interface_ref}"
                _style_matrix_pair(direction_cell, interface_cell, fill=MIRROR_FILL)
            else:
                _style_matrix_pair(direction_cell, interface_cell)


def _style_matrix_pair(direction_cell, interface_cell, fill: str | None = None) -> None:
    for cell in (direction_cell, interface_cell):
        cell.font = Font(name="Arial", size=16 if cell is direction_cell else 11, color=TEXT_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN_GRAY, right=THIN_BLACK if cell is interface_cell else None)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)


def _add_matrix_validations(sheet, count: int) -> None:
    direction_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(DIRECTION_OPTIONS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    interface_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(INTERFACE_OPTIONS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    sheet.add_data_validation(direction_validation)
    sheet.add_data_validation(interface_validation)

    for row_index in range(count):
        row = START_ROW + row_index
        for col_index in range(count):
            if row_index == col_index:
                continue
            direction_col = START_COL + (col_index * 2)
            interface_col = direction_col + 1
            direction_validation.add(sheet.cell(row=row, column=direction_col))
            interface_validation.add(sheet.cell(row=row, column=interface_col))
