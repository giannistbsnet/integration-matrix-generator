import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from integration_matrix.components import Component
from integration_matrix.workbook import (
    DIRECTION_OPTIONS,
    INTERFACE_OPTIONS,
    create_workbook,
    save_workbook,
)


class WorkbookGenerationTest(unittest.TestCase):
    def test_generates_expected_three_component_workbook(self):
        components = [
            Component("ACME", "T1.1", "Order Service"),
            Component("ACME", "T1.2", "Inventory Service"),
            Component("NOVA", "T2.1", "Analytics Engine"),
        ]

        workbook = create_workbook(components)

        self.assertEqual(workbook.sheetnames, ["Integration Matrix", "Descriptions"])

        matrix = workbook["Integration Matrix"]
        descriptions = workbook["Descriptions"]
        self.assertEqual(matrix.max_row, 5)
        self.assertEqual(matrix.max_column, 9)
        self.assertEqual(descriptions.max_row, 4)
        self.assertEqual(descriptions.max_column, 5)

        self.assertEqual(matrix.freeze_panes, "D3")
        self.assertEqual(descriptions.freeze_panes, "A2")
        self.assertIn("D1:E1", {str(range_) for range_ in matrix.merged_cells.ranges})
        self.assertIn("F1:G1", {str(range_) for range_ in matrix.merged_cells.ranges})
        self.assertIn("H1:I1", {str(range_) for range_ in matrix.merged_cells.ranges})

        self.assertEqual(matrix["F3"].value, '=IF(D4="↗","↙",IF(D4="↙","↗",D4))')
        self.assertEqual(matrix["G3"].value, "=E4")
        self.assertEqual(matrix["H3"].value, '=IF(D5="↗","↙",IF(D5="↙","↗",D5))')
        self.assertEqual(matrix["I3"].value, "=E5")
        self.assertEqual(matrix["H4"].value, '=IF(F5="↗","↙",IF(F5="↙","↗",F5))')
        self.assertEqual(matrix["I4"].value, "=G5")

        self.assertTrue(_is_black(matrix["D3"]))
        self.assertTrue(_is_black(matrix["F4"]))
        self.assertTrue(_is_black(matrix["H5"]))

        direction_validation, interface_validation = _validations(matrix)
        self.assertEqual(direction_validation.formula1, f'"{",".join(DIRECTION_OPTIONS)}"')
        self.assertEqual(interface_validation.formula1, f'"{",".join(INTERFACE_OPTIONS)}"')
        self.assertTrue(_contains_validation(direction_validation, "D4"))
        self.assertTrue(_contains_validation(interface_validation, "E4"))

    def test_saves_reopenable_workbook(self):
        components = [Component("ACME", "T1.1", "Order Service")]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "matrix.xlsx"
            save_workbook(create_workbook(components), path)
            reopened = load_workbook(path, data_only=False)

        self.assertEqual(reopened.sheetnames, ["Integration Matrix", "Descriptions"])
        self.assertEqual(reopened["Integration Matrix"]["C3"].value, "Order Service")

    def test_reference_component_count_shape(self):
        components = [
            Component(f"ORG{index:02d}", f"T{index:02d}", f"Component {index:02d}")
            for index in range(1, 29)
        ]

        matrix = create_workbook(components)["Integration Matrix"]

        self.assertEqual(matrix.max_row, 30)
        self.assertEqual(matrix.max_column, 59)
        self.assertEqual(get_column_letter(matrix.max_column), "BG")
        self.assertEqual(matrix["F3"].value, '=IF(D4="↗","↙",IF(D4="↙","↗",D4))')
        self.assertEqual(matrix["BG1"].value, None)
        self.assertEqual(matrix["BF1"].value, "Component 28")
        self.assertTrue(_is_black(matrix["BG30"]))


def _validations(sheet):
    validations = list(sheet.data_validations.dataValidation)
    assert len(validations) == 2
    return validations[0], validations[1]


def _contains_validation(validation, coordinate):
    return any(coordinate in cell_range for cell_range in validation.cells.ranges)


def _is_black(cell):
    color = cell.fill.fgColor.rgb
    return color is not None and color.endswith("000000")


if __name__ == "__main__":
    unittest.main()
